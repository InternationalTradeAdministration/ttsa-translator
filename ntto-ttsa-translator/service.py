# -*- coding: utf-8 -*-
import os
import re
import csv
import tempfile
from io import BytesIO
from azure.storage.blob.blockblobservice import BlockBlobService as bbs
from openpyxl import load_workbook

#Retrieve account specific environment variables
conn_str = os.environ["AzureWebJobsStorage"]
acct_name = re.search('AccountName=(.+?);', conn_str).group(1)
acct_key = re.search('AccountKey=(.+?);', conn_str).group(1)
container_name = os.environ["ContainerName"]

#Handler that converts 
def handler(workbook):
    workbook_name = workbook.name.rsplit("/")[2]
    match_group = re.search(r'[0-9]{4}', workbook.name)
    if not match_group:
        return
    year = match_group.group()
    book = load_workbook(filename=BytesIO(workbook.read()))
    skip_row = 0
    curr_row = 0
    csv_filename = workbook_name.rsplit(".xlsx")[0]+".csv"
    csv_filename = csv_filename.rsplit("/")
    csv_filename.insert(len(csv_filename)-1, "ttsa/translated")
    csv_filename = "/".join(csv_filename)

    temp_file = tempfile.NamedTemporaryFile(mode="r+", delete=False)
    with open(temp_file.name, 'w+', newline='') as nttofile:
        writer = csv.writer(nttofile)
        for sheet in book:
            if sheet.title == "Supply":
                for row in sheet.iter_rows(values_only=True):
                    curr_row = curr_row+1
                    if curr_row != skip_row:
                        print_str = row
                        print_str = list(print_str)
                        to_write = []
                        if "Intermediate" in print_str:
                            if "Commodity" not in print_str:
                                print_str.insert(0, "Commodity")
                            index = print_str.index("Intermediate")
                            print_str.insert(index, "Intermediate Government Expenditures")
                            print_str.insert(index, "Intermediate Private Expenditures")
                            print_str.remove("Intermediate")
                            skip_row = curr_row+1
                            print_str.append("Year")
                            print(print_str)
                        for cell in print_str:
                            if cell is not None:
                                to_write.append(cell)
                        if len(to_write) > 1:
                            if "Commodity" not in to_write:
                                to_write.append(year)
                            writer.writerow(to_write)

    block_blob_service = bbs(account_name=acct_name, account_key=acct_key)
    with open(temp_file.name, 'r+') as upload_data:
        block_blob_service.create_blob_from_text(container_name=container_name, blob_name=csv_filename, text=str(upload_data.read()))

    temp_file.close()
